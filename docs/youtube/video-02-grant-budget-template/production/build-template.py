#!/usr/bin/env python3
"""Builds the free Grant Budget Template workbook (the video's lead magnet).
Tabs: Instructions, Budget, Budget Narrative, Cost Share.
The Budget tab uses real formulas: personnel = salary * % effort, fringe = a
rate * personnel subtotal, indirect = rate * MTDC base (Total Direct minus
Equipment), and Total Project Cost = Total Direct + Indirect. The indirect rate
defaults to the 15% de minimis rate (2024 Uniform Guidance, 2 CFR 200.414).
Brand palette and styling helpers match video-01's build-template.py."""
import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EMERALD = "065F46"; EMERALD_D = "053D2E"; EMERALD_50 = "E7F3EE"
OCHRE = "B9842B"; OCHRE_50 = "F6ECD6"; RED = "B3261E"; RED_50 = "FBE9E7"
PAPER = "FAF7F0"; INK = "17211C"; LINE = "E6E0D2"; MUTED = "6C726A"; WHITE = "FFFFFF"

thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill(): return PatternFill("solid", fgColor=EMERALD_50)
def hf(): return Font(name="Calibri", bold=True, color=EMERALD_D, size=11)
def title_font(sz=16): return Font(name="Calibri", bold=True, color=EMERALD, size=sz)
def body_font(): return Font(name="Calibri", color=INK, size=11)
def muted_font(): return Font(name="Calibri", color=MUTED, size=10, italic=True)
def subtotal_font(): return Font(name="Calibri", bold=True, color=EMERALD_D, size=11)
def subtotal_fill(): return PatternFill("solid", fgColor=EMERALD_50)
def input_fill(): return PatternFill("solid", fgColor=OCHRE_50)

MONEY = '"$"#,##0'
PCT = "0%"

wb = Workbook()

# ---------- Instructions ----------
ins = wb.active; ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 3
ins.column_dimensions["B"].width = 110
def line(r, text, font=None, fill=None):
    c = ins.cell(row=r, column=2, value=text)
    c.font = font or body_font()
    if fill: c.fill = fill
    c.alignment = Alignment(wrap_text=True, vertical="top")
line(2, "Grant Budget Template", title_font(20))
line(3, "A federal-cost-category budget worksheet for nonprofits — by GrantPipe", muted_font())
line(5, "WHAT'S INSIDE", Font(bold=True, color=OCHRE, size=12))
line(6, "• Budget — the standard federal cost categories (Personnel, Fringe, Travel, Equipment, Supplies, Contractual, Other), "
        "then Total Direct Costs, the MTDC base, Indirect, and Total Project Cost. Personnel, fringe, and indirect are formula-driven.")
line(7, "• Budget Narrative — one justification sentence per budget line, so the numbers and the reasons never drift apart.")
line(8, "• Cost Share — record any required match by source, type (Cash / In-Kind), amount, and note.")
line(10, "HOW TO USE IT", Font(bold=True, color=OCHRE, size=12))
line(11, "1. On Budget, replace the sample Personnel rows with your roles. Enter Annual Salary and % Effort; the Cost computes itself.")
line(12, "2. Set the Fringe Rate and Indirect Rate input cells (highlighted). Fringe applies to the personnel subtotal; indirect applies to MTDC.")
line(13, "3. Add Travel, Equipment, Supplies, Contractual, and Other direct costs with a visible basis (quantity x rate).")
line(14, "4. Write a one-line justification for each cost on Budget Narrative. Record any required match on Cost Share.")
line(16, "A NOTE ON THE NUMBERS", Font(bold=True, color=OCHRE, size=12))
line(17, "Equipment is tangible property with a useful life over one year and a per-unit cost at or above the lower of your "
         "capitalization level or $10,000 (raised from $5,000 by the 2024 Uniform Guidance); anything below that is Supplies. The Indirect Rate defaults to the 15% de minimis rate, "
         "which the 2024 Uniform Guidance raised from 10% (2 CFR 200.414). Modified Total Direct Costs (MTDC) exclude equipment and "
         "the portion of each subaward over $50,000; this sample subtracts equipment only. If you have a negotiated indirect rate, "
         "use it instead. This template is a planning tool, not legal or accounting advice.")
line(19, "Built by GrantPipe — donor management & grant compliance for mid-sized nonprofits.  grantpipe.com", Font(color=EMERALD, bold=True, size=11))
line(20, "Free to use and share. Replace the sample rows with your own data.", muted_font())
ins.sheet_properties.tabColor = EMERALD


def style_table(ws, headers, widths, start_row=1):
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = hf(); c.fill = header_fill(); c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row].height = 26


# ---------- Budget ----------
bg = wb.create_sheet("Budget")
bg_headers = ["Category", "Description / Basis", "Quantity", "Rate/Unit", "% Effort", "Cost"]
bg_widths = [22, 38, 11, 14, 11, 16]
style_table(bg, bg_headers, bg_widths, start_row=1)
bg.sheet_view.showGridLines = False

# Column letters: A Category, B Description, C Quantity, D Rate/Unit, E % Effort, F Cost
QTY, RATE, EFFORT, COST = "C", "D", "E", "F"


def cell(row, col, value=None, fmt=None, font=None, fill=None, align="left"):
    c = bg.cell(row=row, column=col, value=value)
    c.border = border
    c.font = font or body_font()
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(col == 2))
    if fill:
        c.fill = fill
    return c


r = 2
# --- Personnel ---
cell(r, 1, "PERSONNEL", font=subtotal_font(), fill=subtotal_fill())
for ci in range(2, 7):
    cell(r, ci, fill=subtotal_fill())
r += 1
personnel_start = r
personnel_rows = [
    ("Project Director", "Oversees program delivery and reporting", None, 95000, 0.25),
    ("Program Coordinator", "Delivers weekly sessions to 40 participants", None, 58000, 0.50),
    ("Data & Compliance Analyst", "Tracks outcomes, prepares funder reports", None, 64000, 0.20),
]
for name, desc, qty, salary, effort in personnel_rows:
    cell(r, 1, name)
    cell(r, 2, desc)
    cell(r, 3, qty, align="right")
    cell(r, 4, salary, fmt=MONEY, align="right")
    cell(r, 5, effort, fmt=PCT, align="right")
    # Cost = Annual Salary (Rate/Unit) * % Effort
    cell(r, 6, f"={RATE}{r}*{EFFORT}{r}", fmt=MONEY, align="right")
    r += 1
personnel_end = r - 1
# Personnel subtotal
cell(r, 1, "Personnel Subtotal", font=subtotal_font(), fill=subtotal_fill())
for ci in (2, 3, 4, 5):
    cell(r, ci, fill=subtotal_fill())
cell(r, 6, f"=SUM({COST}{personnel_start}:{COST}{personnel_end})", fmt=MONEY, font=subtotal_font(), fill=subtotal_fill(), align="right")
personnel_subtotal_row = r
r += 1

# --- Fringe rate input + fringe line ---
cell(r, 1, "Fringe Rate (input)", font=Font(name="Calibri", bold=True, color="7A5410", size=11), fill=input_fill())
cell(r, 2, "Employer payroll taxes, health, retirement — % of personnel", fill=input_fill())
cell(r, 3, None, fill=input_fill())
cell(r, 4, None, fill=input_fill())
fringe_rate_cell = f"{EFFORT}{r}"
cell(r, 5, 0.28, fmt=PCT, font=Font(name="Calibri", bold=True, color="7A5410", size=11), fill=input_fill(), align="right")
cell(r, 6, None, fill=input_fill())
r += 1
cell(r, 1, "Fringe Benefits", font=subtotal_font())
cell(r, 2, "Fringe rate applied to personnel subtotal")
cell(r, 3, None)
cell(r, 4, None)
cell(r, 5, None)
cell(r, 6, f"={COST}{personnel_subtotal_row}*{fringe_rate_cell}", fmt=MONEY, font=subtotal_font(), align="right")
fringe_row = r
r += 1

# --- Travel ---
cell(r, 1, "TRAVEL", font=subtotal_font(), fill=subtotal_fill())
for ci in range(2, 7):
    cell(r, ci, fill=subtotal_fill())
r += 1
travel_start = r
travel_rows = [
    ("Project site visits", "Trips x nights x lodging + mileage", 8, 425),
    ("National grantee convening", "1 traveler: airfare + 3 nights lodging", 1, 1850),
]
for name, desc, qty, rate in travel_rows:
    cell(r, 1, name)
    cell(r, 2, desc)
    cell(r, 3, qty, align="right")
    cell(r, 4, rate, fmt=MONEY, align="right")
    cell(r, 5, None, align="right")
    cell(r, 6, f"={QTY}{r}*{RATE}{r}", fmt=MONEY, align="right")
    r += 1
travel_end = r - 1

# --- Equipment (>= $10,000 item) ---
cell(r, 1, "EQUIPMENT", font=subtotal_font(), fill=subtotal_fill())
for ci in range(2, 7):
    cell(r, ci, fill=subtotal_fill())
r += 1
equipment_start = r
equipment_rows = [
    ("Lab spectrometer", "Unit cost >= $10,000 useful life > 1 yr — excluded from MTDC", 1, 12000),
]
for name, desc, qty, rate in equipment_rows:
    cell(r, 1, name)
    cell(r, 2, desc)
    cell(r, 3, qty, align="right")
    cell(r, 4, rate, fmt=MONEY, align="right")
    cell(r, 5, None, align="right")
    cell(r, 6, f"={QTY}{r}*{RATE}{r}", fmt=MONEY, align="right")
    r += 1
equipment_end = r - 1

# --- Supplies ---
cell(r, 1, "SUPPLIES", font=subtotal_font(), fill=subtotal_fill())
for ci in range(2, 7):
    cell(r, ci, fill=subtotal_fill())
r += 1
supplies_start = r
supplies_rows = [
    ("Participant materials", "Workbooks & kits, 40 participants x $35", 40, 35),
    ("Laptops (under $10,000/unit)", "3 staff laptops at $1,100 — supplies, not equipment", 3, 1100),
]
for name, desc, qty, rate in supplies_rows:
    cell(r, 1, name)
    cell(r, 2, desc)
    cell(r, 3, qty, align="right")
    cell(r, 4, rate, fmt=MONEY, align="right")
    cell(r, 5, None, align="right")
    cell(r, 6, f"={QTY}{r}*{RATE}{r}", fmt=MONEY, align="right")
    r += 1
supplies_end = r - 1

# --- Contractual ---
cell(r, 1, "CONTRACTUAL", font=subtotal_font(), fill=subtotal_fill())
for ci in range(2, 7):
    cell(r, ci, fill=subtotal_fill())
r += 1
contractual_start = r
contractual_rows = [
    ("External evaluator", "Independent outcome evaluation, fixed fee", 1, 12000),
]
for name, desc, qty, rate in contractual_rows:
    cell(r, 1, name)
    cell(r, 2, desc)
    cell(r, 3, qty, align="right")
    cell(r, 4, rate, fmt=MONEY, align="right")
    cell(r, 5, None, align="right")
    cell(r, 6, f"={QTY}{r}*{RATE}{r}", fmt=MONEY, align="right")
    r += 1
contractual_end = r - 1

# --- Other Direct Costs ---
cell(r, 1, "OTHER DIRECT COSTS", font=subtotal_font(), fill=subtotal_fill())
for ci in range(2, 7):
    cell(r, ci, fill=subtotal_fill())
r += 1
other_start = r
other_rows = [
    ("Participant stipends", "40 participants x $50 completion stipend", 40, 50),
    ("Printing & postage", "Outreach and reporting materials", 1, 900),
]
for name, desc, qty, rate in other_rows:
    cell(r, 1, name)
    cell(r, 2, desc)
    cell(r, 3, qty, align="right")
    cell(r, 4, rate, fmt=MONEY, align="right")
    cell(r, 5, None, align="right")
    cell(r, 6, f"={QTY}{r}*{RATE}{r}", fmt=MONEY, align="right")
    r += 1
other_end = r - 1

# --- Total Direct Costs ---
r += 1
# Total Direct = personnel subtotal + fringe + every line item in the direct categories
direct_terms = [
    f"{COST}{personnel_subtotal_row}",
    f"{COST}{fringe_row}",
    f"SUM({COST}{travel_start}:{COST}{travel_end})",
    f"SUM({COST}{equipment_start}:{COST}{equipment_end})",
    f"SUM({COST}{supplies_start}:{COST}{supplies_end})",
    f"SUM({COST}{contractual_start}:{COST}{contractual_end})",
    f"SUM({COST}{other_start}:{COST}{other_end})",
]
cell(r, 1, "TOTAL DIRECT COSTS", font=subtotal_font(), fill=subtotal_fill())
for ci in range(2, 6):
    cell(r, ci, fill=subtotal_fill())
cell(r, 6, "=" + "+".join(direct_terms), fmt=MONEY, font=subtotal_font(), fill=subtotal_fill(), align="right")
total_direct_row = r
r += 1

# --- MTDC base (Total Direct minus Equipment) ---
cell(r, 1, "MTDC Base", font=subtotal_font())
cell(r, 2, "Modified Total Direct Costs = Total Direct minus Equipment")
cell(r, 3, None)
cell(r, 4, None)
cell(r, 5, None)
cell(r, 6, f"={COST}{total_direct_row}-SUM({COST}{equipment_start}:{COST}{equipment_end})", fmt=MONEY, font=subtotal_font(), align="right")
mtdc_row = r
r += 1

# --- Indirect rate input ---
cell(r, 1, "Indirect Rate (input)", font=Font(name="Calibri", bold=True, color="7A5410", size=11), fill=input_fill())
cell(r, 2, "De minimis rate = 15% of MTDC (2024 Uniform Guidance, 2 CFR 200.414)", fill=input_fill())
cell(r, 3, None, fill=input_fill())
cell(r, 4, None, fill=input_fill())
indirect_rate_cell = f"{EFFORT}{r}"
cell(r, 5, 0.15, fmt=PCT, font=Font(name="Calibri", bold=True, color="7A5410", size=11), fill=input_fill(), align="right")
cell(r, 6, None, fill=input_fill())
r += 1

# --- Indirect Costs ---
cell(r, 1, "INDIRECT COSTS", font=subtotal_font(), fill=subtotal_fill())
cell(r, 2, "Indirect Rate applied to the MTDC base", fill=subtotal_fill())
cell(r, 3, None, fill=subtotal_fill())
cell(r, 4, None, fill=subtotal_fill())
cell(r, 5, None, fill=subtotal_fill())
cell(r, 6, f"={COST}{mtdc_row}*{indirect_rate_cell}", fmt=MONEY, font=subtotal_font(), fill=subtotal_fill(), align="right")
indirect_row = r
r += 1

# --- Total Project Cost ---
cell(r, 1, "TOTAL PROJECT COST", font=Font(name="Calibri", bold=True, color=WHITE, size=12),
     fill=PatternFill("solid", fgColor=EMERALD))
for ci in range(2, 6):
    cell(r, ci, fill=PatternFill("solid", fgColor=EMERALD))
cell(r, 6, f"={COST}{total_direct_row}+{COST}{indirect_row}", fmt=MONEY,
     font=Font(name="Calibri", bold=True, color=WHITE, size=12),
     fill=PatternFill("solid", fgColor=EMERALD), align="right")
total_project_row = r

bg.freeze_panes = "A2"
bg.sheet_properties.tabColor = EMERALD


# ---------- Budget Narrative ----------
nb = wb.create_sheet("Budget Narrative")
nb_headers = ["Category", "Line Item", "Justification"]
nb_widths = [22, 30, 80]
style_table(nb, nb_headers, nb_widths, start_row=1)
narrative_rows = [
    ("Personnel", "Project Director", "0.25 FTE to supervise program delivery, manage funder relationships, and certify time and effort across the award period."),
    ("Personnel", "Program Coordinator", "0.50 FTE to deliver weekly sessions to 40 participants and maintain attendance and outcome records."),
    ("Personnel", "Data & Compliance Analyst", "0.20 FTE to track outcomes, reconcile spending to budget, and prepare funder financial and narrative reports."),
    ("Fringe Benefits", "Fringe", "Employer payroll taxes, health insurance, and retirement at the organization's audited fringe rate applied to charged salaries."),
    ("Travel", "Project site visits", "Eight monitoring visits to program sites at the federal per-night lodging and mileage rates."),
    ("Travel", "National grantee convening", "One staff member to the required annual grantee convening: airfare plus three nights lodging at published rates."),
    ("Equipment", "Lab spectrometer", "One spectrometer at $12,000, useful life over one year; at or above the $10,000 threshold, so capitalized and excluded from the MTDC indirect base."),
    ("Supplies", "Participant materials", "Workbooks and activity kits for 40 enrolled participants at $35 each."),
    ("Supplies", "Laptops (under $10,000/unit)", "Three staff laptops at $1,100 each; below the $10,000 capitalization threshold, so budgeted as supplies."),
    ("Contractual", "External evaluator", "Independent third-party evaluation of program outcomes under a fixed-fee agreement."),
    ("Other Direct Costs", "Participant stipends", "Completion stipends of $50 for 40 participants to support retention through the full program."),
    ("Other Direct Costs", "Printing & postage", "Outreach, enrollment, and reporting materials over the project period."),
    ("Indirect Costs", "De minimis indirect", "15% de minimis rate applied to Modified Total Direct Costs, per 2 CFR 200.414, in the absence of a negotiated rate."),
]
for ri, (cat, item, just) in enumerate(narrative_rows, start=2):
    c1 = nb.cell(row=ri, column=1, value=cat); c1.border = border; c1.font = body_font(); c1.alignment = Alignment(vertical="top")
    c2 = nb.cell(row=ri, column=2, value=item); c2.border = border; c2.font = body_font(); c2.alignment = Alignment(vertical="top")
    c3 = nb.cell(row=ri, column=3, value=just); c3.border = border; c3.font = body_font(); c3.alignment = Alignment(vertical="top", wrap_text=True)
nb.freeze_panes = "A2"
nb.sheet_properties.tabColor = OCHRE


# ---------- Cost Share ----------
cs = wb.create_sheet("Cost Share")
cs_headers = ["Source", "Type (Cash/In-Kind)", "Amount", "Notes"]
cs_widths = [28, 20, 16, 50]
style_table(cs, cs_headers, cs_widths, start_row=1)
costshare_rows = [
    ("Organization unrestricted funds", "Cash", 10000, "Board-approved match committed at application."),
    ("Volunteer mentor hours", "In-Kind", 6000, "120 hours valued at the independent-sector volunteer rate."),
    ("Donated meeting space", "In-Kind", 3000, "Partner-provided space for participant sessions, valued at fair rental value."),
]
for ri, (src, typ, amt, note) in enumerate(costshare_rows, start=2):
    c1 = cs.cell(row=ri, column=1, value=src); c1.border = border; c1.font = body_font(); c1.alignment = Alignment(vertical="center")
    c2 = cs.cell(row=ri, column=2, value=typ); c2.border = border; c2.font = body_font(); c2.alignment = Alignment(vertical="center")
    c3 = cs.cell(row=ri, column=3, value=amt); c3.border = border; c3.font = body_font(); c3.number_format = MONEY; c3.alignment = Alignment(horizontal="right", vertical="center")
    c4 = cs.cell(row=ri, column=4, value=note); c4.border = border; c4.font = body_font(); c4.alignment = Alignment(vertical="center", wrap_text=True)
# Total match row
tr = 2 + len(costshare_rows)
tc1 = cs.cell(row=tr, column=1, value="Total Match"); tc1.border = border; tc1.font = subtotal_font(); tc1.fill = subtotal_fill()
tc2 = cs.cell(row=tr, column=2); tc2.border = border; tc2.fill = subtotal_fill()
tc3 = cs.cell(row=tr, column=3, value=f"=SUM(C2:C{tr-1})"); tc3.border = border; tc3.font = subtotal_font(); tc3.fill = subtotal_fill(); tc3.number_format = MONEY; tc3.alignment = Alignment(horizontal="right", vertical="center")
tc4 = cs.cell(row=tr, column=4); tc4.border = border; tc4.fill = subtotal_fill()
cs.freeze_panes = "A2"
cs.sheet_properties.tabColor = OCHRE


# ---------- Save ----------
production_dir = os.path.dirname(__file__)
assets_dir = os.path.join(production_dir, "..", "assets")
os.makedirs(assets_dir, exist_ok=True)
asset_out = os.path.join(assets_dir, "GrantPipe-Grant-Budget-Template.xlsx")
wb.save(asset_out)
print("wrote", os.path.relpath(asset_out))

# Delivery copy used as the email-gated lead magnet asset.
delivery_out = os.path.join(
    production_dir, "..", "..", "..", "..",
    "apps", "site", "src", "assets", "lead-magnets", "grant-budget-template.xlsx",
)
delivery_out = os.path.normpath(delivery_out)
os.makedirs(os.path.dirname(delivery_out), exist_ok=True)
shutil.copyfile(asset_out, delivery_out)
print("wrote", os.path.relpath(delivery_out, production_dir))
