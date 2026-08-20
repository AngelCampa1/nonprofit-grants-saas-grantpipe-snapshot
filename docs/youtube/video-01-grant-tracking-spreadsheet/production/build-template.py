#!/usr/bin/env python3
"""Builds the free Grant Tracking Spreadsheet template (the video's lead magnet).
Tabs: Instructions, Grant Register, Budget vs Actual, Expense Log, Reporting.
Real formulas (SUMIFS / arithmetic / COUNTIFS), dropdown validation, and
conditional formatting — usable as-is and matched to the video walkthrough."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

EMERALD = "065F46"; EMERALD_D = "053D2E"; EMERALD_50 = "E7F3EE"
OCHRE = "B9842B"; OCHRE_50 = "F6ECD6"; RED = "B3261E"; RED_50 = "FBE9E7"
PAPER = "FAF7F0"; INK = "17211C"; LINE = "E6E0D2"; MUTED = "6C726A"; WHITE = "FFFFFF"

thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def header_fill(): return PatternFill("solid", fgColor=EMERALD_50)
def hf(): return Font(name="Calibri", bold=True, color=EMERALD_D, size=11)
def title_font(sz=16): return Font(name="Calibri", bold=True, color=EMERALD, size=sz)
def body_font(): return Font(name="Calibri", color=INK, size=11)
def muted_font(): return Font(name="Calibri", color=MUTED, size=10, italic=True)

wb = Workbook()

# ---------- Instructions ----------
ins = wb.active; ins.title = "Instructions"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 3
ins.column_dimensions["B"].width = 110
def line(r, text, font=None, fill=None):
    c = ins.cell(row=r, column=2, value=text)
    c.font = font or body_font()
    if fill: c.fill = fill
    c.alignment = Alignment(wrap_text=True, vertical="top")
line(2, "Grant Tracking Spreadsheet", title_font(20))
line(3, "A compliance-aware template for nonprofits — by GrantPipe", muted_font())
line(5, "WHAT'S INSIDE", Font(bold=True, color=OCHRE, size=12))
line(6, "• Grant Register — one row per grant, with the restricted-funds and period-of-performance columns most trackers skip.")
line(7, "• Budget vs Actual — track spend against each grant's budget, category by category. 'Spent' fills itself from the Expense Log.")
line(8, "• Expense Log — one row per dollar charged to a grant. The Grant ID ties each expense back to its budget line.")
line(9, "• Reporting — funder report deadlines, with colour warnings as due dates approach.")
line(11, "HOW TO USE IT", Font(bold=True, color=OCHRE, size=12))
line(12, "1. List every grant on the Grant Register. Give each a short Grant ID (e.g. ED-2026-01) and mark whether it's Restricted.")
line(13, "2. On Budget vs Actual, enter the funder-approved budget by category for each Grant ID.")
line(14, "3. Log every expense on the Expense Log as it happens. 'Spent' and 'Remaining' update automatically.")
line(15, "4. Add funder report deadlines on Reporting. Rows turn amber within 30 days and red once overdue.")
line(17, "A NOTE ON COMPLIANCE", Font(bold=True, color=OCHRE, size=12))
line(18, "Restricted grant dollars are tied to a specific purpose. Keep them separate from general operating funds. This template helps "
         "you see your position at a glance — it is not legal or accounting advice. When you outgrow a spreadsheet (multiple editors, "
         "an audit trail, split-fund allocations), that's the signal to move to dedicated software.")
line(20, "Built by GrantPipe — donor management & grant compliance for mid-sized nonprofits.  grantpipe.com", Font(color=EMERALD, bold=True, size=11))
line(21, "Free to use and share. Replace the sample rows with your own data.", muted_font())
ins.sheet_properties.tabColor = EMERALD

def style_table(ws, headers, widths, start_row=1):
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = w
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = hf(); c.fill = header_fill(); c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row].height = 26

def fill_rows(ws, data, start_row, ncols, money_cols=(), date_cols=()):
    for ri, row in enumerate(data, start=start_row):
        for ci in range(1, ncols + 1):
            c = ws.cell(row=ri, column=ci, value=row[ci-1] if ci-1 < len(row) else None)
            c.border = border; c.font = body_font()
            c.alignment = Alignment(vertical="center")
            if ci in money_cols:
                c.number_format = '"$"#,##0'; c.alignment = Alignment(horizontal="right", vertical="center")
            if ci in date_cols:
                c.number_format = "mm/dd/yyyy"

# ---------- Grant Register ----------
gr = wb.create_sheet("Grant Register")
gr_headers = ["Grant ID","Grant Name","Funder","Funder Type","Award Amount","Start Date","End Date","Restricted?","Restriction Purpose","Status"]
gr_widths  = [14,26,22,16,16,13,13,13,30,14]
# Dashboard strip above the table
gr.sheet_view.showGridLines = False
dash = [("Total Awarded",'=SUM(E4:E200)','"$"#,##0'),
        ("Total Spent",'=SUM(\'Budget vs Actual\'!D4:D400)','"$"#,##0'),
        ("% Restricted",'=IF(SUM(E4:E200)=0,0,SUMIF(H4:H200,"Yes",E4:E200)/SUM(E4:E200))','0%'),
        ("Reports Due / 30d",'=COUNTIFS(Reporting!C4:C200,"<="&(TODAY()+30),Reporting!C4:C200,">="&TODAY(),Reporting!D4:D200,"")','0')]
# 4 KPI cells across the top of the register (label row 1, value row 2)
for (label,formula,fmt),(c1,c2) in zip(dash,[(1,2),(3,4),(5,6),(7,8)]):
    lc = gr.cell(row=1,column=c1,value=label); lc.font=Font(bold=True,color=MUTED,size=9)
    vc = gr.cell(row=2,column=c1,value=formula); vc.font=Font(bold=True,color=EMERALD,size=16); vc.number_format=fmt
gr.row_dimensions[2].height = 22
style_table(gr, gr_headers, gr_widths, start_row=3)
gr_data = [
    ["ED-2026-01","After-School STEM","Dept. of Education","Federal",50000,"2025-10-01","2026-09-30","Yes","STEM program only","Active"],
    ["HF-2026-02","Family Literacy","Hartwell Foundation","Foundation",28000,"2026-01-01","2026-12-31","Yes","Literacy materials & staff","Active"],
    ["ST-2026-03","Summer Meals","State DHS","State",41500,"2026-05-01","2026-08-31","Yes","Meal program only","Awarded"],
    ["GEN-OPS-00","General Operating","Community Fund","Foundation",15000,"2026-01-01","2026-12-31","No","—","Active"],
]
# convert date strings to dates
import datetime
def d(s): return datetime.datetime.strptime(s,"%Y-%m-%d")
for r in gr_data:
    r[5]=d(r[5]); r[6]=d(r[6])
fill_rows(gr, gr_data, start_row=4, ncols=10, money_cols=(5,), date_cols=(6,7))
# Validation: Funder Type, Restricted, Status
dv_type = DataValidation(type="list", formula1='"Federal,State,Local,Foundation,Corporate,Association,Other"', allow_blank=True)
dv_restr = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
dv_status = DataValidation(type="list", formula1='"Applied,Awarded,Active,Reporting,Closed"', allow_blank=True)
gr.add_data_validation(dv_type); gr.add_data_validation(dv_restr); gr.add_data_validation(dv_status)
dv_type.add("D4:D200"); dv_restr.add("H4:H200"); dv_status.add("J4:J200")
# Conditional format: highlight Restricted=Yes rows' Restricted cell ochre
gr.conditional_formatting.add("H4:H200", CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid",fgColor=OCHRE_50), font=Font(color="7A5410",bold=True)))
gr.freeze_panes = "A4"
gr.sheet_properties.tabColor = EMERALD

# ---------- Budget vs Actual ----------
ba = wb.create_sheet("Budget vs Actual")
ba_headers = ["Grant ID","Budget Category","Budgeted","Spent (auto)","Remaining","% Spent"]
ba_widths  = [14,28,15,16,15,11]
style_table(ba, ba_headers, ba_widths, start_row=1)
ba_rows = [
    ["ED-2026-01","Personnel",32000],
    ["ED-2026-01","Supplies",6000],
    ["ED-2026-01","Travel",8000],
    ["ED-2026-01","Indirect (10%)",4000],
    ["HF-2026-02","Personnel",18000],
    ["HF-2026-02","Materials",10000],
]
for ri,row in enumerate(ba_rows, start=2):
    gid,cat,bud = row
    ba.cell(row=ri,column=1,value=gid).border=border
    ba.cell(row=ri,column=2,value=cat).border=border
    bc=ba.cell(row=ri,column=3,value=bud); bc.number_format='"$"#,##0'; bc.border=border; bc.alignment=Alignment(horizontal="right")
    # Spent = SUMIFS(ExpenseLog Amount, GrantID, this, Category, this)
    sc=ba.cell(row=ri,column=4,value=f'=SUMIFS(\'Expense Log\'!$E$4:$E$1000,\'Expense Log\'!$B$4:$B$1000,A{ri},\'Expense Log\'!$C$4:$C$1000,B{ri})')
    sc.number_format='"$"#,##0'; sc.border=border; sc.alignment=Alignment(horizontal="right")
    rc=ba.cell(row=ri,column=5,value=f'=C{ri}-D{ri}'); rc.number_format='"$"#,##0'; rc.border=border; rc.alignment=Alignment(horizontal="right")
    pc=ba.cell(row=ri,column=6,value=f'=IF(C{ri}=0,0,D{ri}/C{ri})'); pc.number_format='0%'; pc.border=border; pc.alignment=Alignment(horizontal="right")
    for ci in range(1,7): ba.cell(row=ri,column=ci).font=body_font()
ba.conditional_formatting.add(f"E2:E400", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid",fgColor=RED_50), font=Font(color=RED,bold=True)))
ba.conditional_formatting.add(f"F2:F400", CellIsRule(operator="greaterThan", formula=["1"], fill=PatternFill("solid",fgColor=RED_50), font=Font(color=RED,bold=True)))
dv_gid_ba = DataValidation(type="list", formula1="='Grant Register'!$A$4:$A$200", allow_blank=True)
ba.add_data_validation(dv_gid_ba); dv_gid_ba.add("A2:A400")
ba.freeze_panes="A2"; ba.sheet_properties.tabColor=OCHRE

# ---------- Expense Log ----------
el = wb.create_sheet("Expense Log")
el_headers=["Date","Grant ID","Category","Vendor","Amount","Notes"]
el_widths=[13,14,22,26,14,30]
style_table(el, el_headers, el_widths, start_row=3)
el.cell(row=1,column=1,value="Log every dollar charged to a grant. 'Grant ID' + 'Category' must match the Budget vs Actual rows so SUMIFS can find them.").font=muted_font()
el_rows=[
    [d("2026-03-04"),"ED-2026-01","Supplies","Lakeshore Learning",1240,"Lab kits"],
    [d("2026-03-09"),"ED-2026-01","Travel","Delta Air Lines",612,"Site visit"],
    [d("2026-03-12"),"HF-2026-02","Personnel","Payroll",3800,"Tutor wages"],
    [d("2026-03-18"),"ED-2026-01","Personnel","Payroll",4100,"Instructor"],
    [d("2026-03-22"),"ED-2026-01","Travel","Marriott",980,"Conference"],
]
fill_rows(el, el_rows, start_row=4, ncols=6, money_cols=(5,), date_cols=(1,))
dv_gid_el=DataValidation(type="list", formula1="='Grant Register'!$A$4:$A$200", allow_blank=True)
el.add_data_validation(dv_gid_el); dv_gid_el.add("B4:B1000")
el.freeze_panes="A4"; el.sheet_properties.tabColor=EMERALD

# ---------- Reporting ----------
rp = wb.create_sheet("Reporting")
rp_headers=["Grant ID","Report Type","Due Date","Submitted Date","Status (auto)"]
rp_widths=[14,26,14,16,20]
style_table(rp, rp_headers, rp_widths, start_row=3)
rp.cell(row=1,column=1,value="Due dates turn amber within 30 days and red once overdue and unsubmitted.").font=muted_font()
rp_rows=[
    ["ED-2026-01","Quarterly Financial",d("2026-06-15"),None],
    ["HF-2026-02","Interim Narrative",d("2026-05-01"),None],
    ["ST-2026-03","Final Report",d("2026-09-30"),None],
]
for ri,row in enumerate(rp_rows, start=4):
    rp.cell(row=ri,column=1,value=row[0]).border=border
    rp.cell(row=ri,column=2,value=row[1]).border=border
    dc=rp.cell(row=ri,column=3,value=row[2]); dc.number_format="mm/dd/yyyy"; dc.border=border
    sc=rp.cell(row=ri,column=4,value=row[3]); sc.number_format="mm/dd/yyyy"; sc.border=border
    st=rp.cell(row=ri,column=5,value=f'=IF(D{ri}<>"","Submitted",IF(C{ri}<TODAY(),"OVERDUE",IF(C{ri}<=TODAY()+30,"Due soon","Upcoming")))')
    st.border=border
    for ci in range(1,6): rp.cell(row=ri,column=ci).font=body_font()
# conditional format due date col C
rp.conditional_formatting.add("C4:C200", FormulaRule(formula=['AND($C4<>"",$D4="",$C4<TODAY())'], fill=PatternFill("solid",fgColor=RED_50), font=Font(color=RED,bold=True)))
rp.conditional_formatting.add("C4:C200", FormulaRule(formula=['AND($C4<>"",$D4="",$C4>=TODAY(),$C4<=TODAY()+30)'], fill=PatternFill("solid",fgColor=OCHRE_50), font=Font(color="7A5410",bold=True)))
dv_gid_rp=DataValidation(type="list", formula1="='Grant Register'!$A$4:$A$200", allow_blank=True)
rp.add_data_validation(dv_gid_rp); dv_gid_rp.add("A4:A200")
rp.freeze_panes="A4"; rp.sheet_properties.tabColor=OCHRE

out_dir = os.path.join(os.path.dirname(__file__), "..", "assets")
os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir, "GrantPipe-Grant-Tracking-Template.xlsx")
wb.save(out)
print("wrote", os.path.relpath(out))
